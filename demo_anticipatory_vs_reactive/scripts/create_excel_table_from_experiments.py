from openpyxl import Workbook
import openpyxl


import os
wb = Workbook()
ws = wb.active
import csv

ws.append(["User_Id", "Predictive/Reactive", "Task", "Threshold Plan", "Threshold Execute", "Gaze object", 
           "Time (s): Gaze dec. prob> threshold plan", "Time (s): Gaze dec. prob> threshold execute", 
           "Time (s): Compute Grasp Poses", "Time (s): Feasible Pose", "Time (s): Plan trajectory reaching", 
           "Time (s): Start trajectory execution", "Asr command to grasp", "Time (s): Asr command to grasp", 
           "Time(s): Robot in reaching pose", "Time(s): Object grasped", "Time(s): Object up", "Time(s): Object delivered"])

general_directory = "../experimentos"
users_ids = []
for id in range(3,23):
    if os.path.isdir(general_directory+"/user_"+str(id)):
        users_ids.append("user_"+str(id))
print(users_ids)     
for str_predictive in ["Predictive", "Reactive"]:
    for str_user_id in users_ids:
        for str_task in os.listdir(general_directory+"/"+str_user_id+"/"+str_predictive):
            print(str_task)
            user_id = int(str_user_id.split("_")[1])
            str_gaze_object = None
            time_gaze_dec_prob_threshold_plan= None
            time_gaze_dec_prob_threshold_execute = None
            time_compute_grasp_poses = None
            time_feasible_pose = None
            time_plan_reaching_pose = None
            time_start_trajectory_execution = None
            str_asr_object = None
            time_asr_object = None
            time_robot_in_reaching_pose = None
            time_object_grasped = None
            time_object_up = None
            time_object_delivered = None
            already_exists = False

            with open(general_directory+"/"+str_user_id+"/"+str_predictive+"/"+str_task, newline='', ) as csvfile:
                spamreader = csv.reader(csvfile, delimiter=',')
                for row in spamreader:
                    print(row)
                    if(row[0] == "Gaze object"):
                        str_gaze_object = row[1]
                        if (row[5] == "Threshold plan"):
                            for i in range(5,len(row)):
                                if(row[i] == "Seconds from demo start time"):
                                    time_gaze_dec_prob_threshold_plan = row[i+1]
                                    already_exists = True
                        if (row[5] == "Threshold execute"):
                            for i in range(5,len(row)):
                                if(row[i] == "Seconds from demo start time"):
                                    time_gaze_dec_prob_threshold_execute = row[i+1]
                                    if not already_exists:
                                        time_gaze_dec_prob_threshold_plan = time_gaze_dec_prob_threshold_execute
                    if(row[0] == "Compute grasp poses object"):
                        for i in range(5,len(row)):
                            if(row[i] == "Seconds from demo start time"):
                                time_compute_grasp_poses = row[i+1]
                    if(row[0] == "Feasible reaching pose for object"):
                        for i in range(5,len(row)):
                            if(row[i] == "Seconds from demo start time"):
                                time_feasible_pose = row[i+1]
                    if(row[0] == "Plan trajectory reaching pose for object"):
                        for i in range(5,len(row)):
                            if(row[i] == "Seconds from demo start time"):
                                time_plan_reaching_pose = row[i+1]
                    if(row[0] == "Start Execution trajectory to reach"):
                        for i in range(5, len(row)):
                            if (row[i] == "Seconds from demo start time"):
                                time_start_trajectory_execution = row[i + 1]
                    if(row[0] == "Asr command to grasp object"):
                        str_asr_object = row[1]
                        for i in range(5, len(row)):
                            if (row[i] == "Seconds from demo start time"):
                                time_asr_object = row[i + 1]
                    if(row[0] == "Robot in reaching Pose trajectory"):
                        for i in range(5, len(row)):
                            if (row[i] == "Seconds from demo start time"):
                                time_robot_in_reaching_pose = row[i + 1]
                    if(row[0] == "Object grasped"):
                        for i in range(5, len(row)):
                            if (row[i] == "Seconds from demo start time"):
                                time_object_grasped = row[i + 1]
                    if(row[0] == "Object up"):
                        for i in range(5, len(row)):
                            if (row[i] == "Seconds from demo start time"):
                                time_object_up = row[i + 1]
                    if(row[0] == "Object delivered"):
                        for i in range(5, len(row)):
                            if (row[i] == "Seconds from demo start time"):
                                time_object_delivered = row[i + 1]

                
                data =[user_id, str_predictive, 2, 0.2, 0.5, str_gaze_object, time_gaze_dec_prob_threshold_plan,
                        time_gaze_dec_prob_threshold_execute, time_compute_grasp_poses, time_feasible_pose, time_plan_reaching_pose,
                        time_start_trajectory_execution, str_asr_object, time_asr_object, time_robot_in_reaching_pose,time_object_grasped, time_object_up, time_object_delivered]
                print(data)
                ws.append(data)


# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("OrganizacionExperimentosSHARON.xlsx")

# Define variable to read the active sheet:
worksheet = wookbook.active

predictive_all_tests = []
reactive_all_tests = []
# Iterate the loop to read the cell values
for i in range(25, 45):
    for col in worksheet.iter_cols(5,5):
        predictive = [col[i].value]
    for col in worksheet.iter_cols(6,6):
        reactive = [col[i].value]
    res = [eval(i) for i in predictive]
    print(res)
    res_reactive = [eval(i) for i in reactive]

    predictive_all_tests.append(res[0][0])
    predictive_all_tests.append(res[0][1])
    predictive_all_tests.append(res[0][2])

    # reactive_all_tests.append(res_reactive[0][0])
    # reactive_all_tests.append(res_reactive[0][1])
    # reactive_all_tests.append(res_reactive[0][2])


print(predictive_all_tests)
for i in range(len(predictive_all_tests)):
    print(predictive_all_tests[i])
    ws.cell(row=2+i, column=3, value=predictive_all_tests[i])
# for i in range(0, len(predictive_all_tests)):
#     ws.cell(row=22+i, column=4, value=reactive_all_tests[i])

wb.save("table.xlsx")